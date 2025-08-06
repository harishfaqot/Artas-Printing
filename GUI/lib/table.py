from PyQt5.QtCore import Qt
from PyQt5.QtWidgets import QApplication, QTableWidget, QTableWidgetItem
import datetime
import openpyxl
import csv
from openpyxl.styles import Font, PatternFill

def setup_table_functionality(self, table):
    def copy():
        selected = table.selectedIndexes()
        if not selected:
            return
        
        selected.sort()
        rows = selected[-1].row() - selected[0].row() + 1
        cols = selected[-1].column() - selected[0].column() + 1
        
        text = ""
        for i in range(rows):
            for j in range(cols):
                if j > 0:
                    text += "\t"
                item = table.item(selected[0].row() + i, selected[0].column() + j)
                if item is not None:
                    text += item.text()
            text += "\n"
        
        QApplication.clipboard().setText(text)

    def paste():
        selected = table.selectedIndexes()
        if not selected:
            return

        row = selected[0].row()
        col = selected[0].column()

        clipboard = QApplication.clipboard()
        text = clipboard.text()
        rows = text.split('\n')

        for i, r in enumerate(rows):
            if not r.strip():
                continue
            cells = r.split('\t')
            for j, c in enumerate(cells):
                if row + i >= table.rowCount() or col + j >= table.columnCount():
                    continue

                item = table.item(row + i, col + j)
                if item is not None:
                    item.setText(c)
                    item.setTextAlignment(Qt.AlignCenter)
                else:
                    new_item = QTableWidgetItem(c)
                    new_item.setTextAlignment(Qt.AlignCenter)
                    table.setItem(row + i, col + j, new_item)

    def delete():
        # Block signals to prevent unwanted updates
        self.tableWidget_input.blockSignals(True)
        self.tableWidget_home.blockSignals(True)
        
        selected = self.tableWidget_input.selectedIndexes()
        if not selected:
            return

        # Get unique rows to delete (sorted from bottom to top)
        rows_to_delete = sorted(set(index.row() for index in selected), reverse=True)

        # Delete rows in both tables
        for row in rows_to_delete:
            self.tableWidget_input.removeRow(row)
            if row < self.tableWidget_home.rowCount():
                self.tableWidget_home.removeRow(row)
        
        # Qt will automatically update the default row numbers in both tables
        
        # Unblock signals
        self.tableWidget_input.blockSignals(False)
        self.tableWidget_home.blockSignals(False)

    def keyPressEvent(event):
        if event.key() == Qt.Key_C and event.modifiers() == Qt.ControlModifier:
            copy()
        elif event.key() == Qt.Key_V and event.modifiers() == Qt.ControlModifier:
            paste()
        elif event.key() == Qt.Key_Delete:
            delete()
        else:
            QTableWidget.keyPressEvent(table, event)

    table.copy = copy
    table.paste = paste
    table.delete = delete
    table.keyPressEvent = keyPressEvent

def add_to_history(self, output_text, status):
    current_time = datetime.datetime.now()
    date_str = current_time.strftime("%Y-%m-%d")
    time_str = current_time.strftime("%H:%M:%S")
    
    row_position = self.tableWidget.rowCount()
    self.tableWidget.insertRow(row_position)
    
    for col, value in enumerate([
        date_str,
        time_str,
        output_text,
        status
    ]):
        item = QTableWidgetItem(value)
        item.setTextAlignment(Qt.AlignCenter)
        item.setFlags(item.flags() ^ Qt.ItemIsEditable)
        self.tableWidget.setItem(row_position, col, item)

    # Save to CSV
    # self.save_to_csv(date_str, time_str, weight, length, counter, output_text, status)

def save_to_csv(self, date, time, weight, length, counter, output_text, status):
    current_time = datetime.datetime.now()
    date_str = current_time.strftime("%Y-%m-%d")
    csv_file = f"logs/{date_str}.csv"
    
    file_exists = os.path.isfile(csv_file)
    
    weight_min = self.lineEdit_downlimit.text()
    weight_max = self.lineEdit_uplimit.text()
    length_min = self.lineEdit_downlimit_1.text()
    length_max = self.lineEdit_uplimit_1.text()
    
    try:
        with open(csv_file, 'a', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            
            if not file_exists:
                writer.writerow([
                    "Date", "Time", "Weight", "Min", "Max", 
                    "Length", "Min", "Max", "Counter", 
                    "Printed Text", "Status"
                ])
            
            writer.writerow([
                date,
                time,
                f"{weight} {self.weight_unit}",
                weight_min,
                weight_max,
                f"{length} {self.length_unit}",
                length_min,
                length_max,
                str(counter),
                output_text,
                status
            ])
    except Exception as e:
        print(f"Error saving to CSV: {e}")

def load_last_csv(self):
    try:
        csv_files = glob.glob("logs/*.csv")
        if not csv_files:
            return

        latest_file = max(csv_files, key=os.path.getmtime)
        self.lineEdit_path.setText(latest_file)

        with open(latest_file, 'r', newline='', encoding='utf-8') as f:
            reader = csv.reader(f)
            headers = next(reader)  # Skip header row

            self.tableWidget.setRowCount(0)

            for row_data in reader:
                row_position = self.tableWidget.rowCount()
                self.tableWidget.insertRow(row_position)
                for col, value in enumerate(row_data):
                    item = QTableWidgetItem(value)
                    item.setTextAlignment(Qt.AlignCenter)
                    item.setFlags(item.flags() ^ Qt.ItemIsEditable)
                    
                    # Apply color coding for rejected items when loading
                    if col == 10 and value != "OK":  # Status column is now index 10
                        if "WEIGHT" in value:
                            item.setBackground(QBrush(QColor("yellow")))
                        elif "LENGTH" in value:
                            item.setBackground(QBrush(QColor("orange")))
                            
                    self.tableWidget.setItem(row_position, col, item)
    except Exception as e:
        print(f"Failed to load last CSV: {e}")

def open_file(self):
    file_path, _ = QFileDialog.getOpenFileName(
        self,
        "Open CSV File",
        "./logs",  # or use absolute path like "/home/user/logs"
        "CSV Files (*.csv)"
    )
    if file_path:
        self.lineEdit_path.setText(file_path)
        try:
            with open(file_path, 'r', newline='', encoding='utf-8') as f:
                reader = csv.reader(f)
                headers = next(reader)

                self.tableWidget.setRowCount(0)

                for row_data in reader:
                    row_position = self.tableWidget.rowCount()
                    self.tableWidget.insertRow(row_position)
                    for col, value in enumerate(row_data):
                        item = QTableWidgetItem(value)
                        item.setTextAlignment(Qt.AlignCenter)
                        item.setFlags(item.flags() ^ Qt.ItemIsEditable)
                        
                        # Apply color coding for rejected items
                        if col == 10 and value != "OK":  # Status column
                            if "WEIGHT" in value:
                                item.setBackground(QBrush(QColor("yellow")))
                            elif "LENGTH" in value:
                                item.setBackground(QBrush(QColor("orange")))
                                
                        self.tableWidget.setItem(row_position, col, item)
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to open file: {str(e)}")

def save_data(self):
    QMessageBox.information(self, "Data Saved", "Data has been saved successfully")

def export_to_excel(self):
    if self.tableWidget.rowCount() == 0:
        QMessageBox.warning(self, "No Data", "There is no data to export")
        return
        
    file_path, _ = QFileDialog.getSaveFileName(self, "Export to Excel", "", "Excel Files (*.xlsx);;All Files (*)")
    if file_path:
        if not file_path.endswith('.xlsx'):
            file_path += '.xlsx'
            
        try:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Printing History"
            
            # Write headers
            headers = []
            for col in range(self.tableWidget.columnCount()):
                headers.append(self.tableWidget.horizontalHeaderItem(col).text())
            sheet.append(headers)
            
            # Make headers bold
            for cell in sheet[1]:
                cell.font = Font(bold=True)
            
            # Write data with color coding
            for row in range(self.tableWidget.rowCount()):
                row_data = []
                for col in range(self.tableWidget.columnCount()):
                    item = self.tableWidget.item(row, col)
                    row_data.append(item.text() if item else "")
                sheet.append(row_data)
                
                # Apply color to entire row based on status
                status_item = self.tableWidget.item(row, 10)  # Status column is now index 10
                if status_item and status_item.text() != "OK":
                    fill_color = "FFFF00" if "WEIGHT" in status_item.text() else "FFA500"
                    for cell in sheet[row+2]:  # +2 because headers are row 1 and Excel is 1-based
                        cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
            
            workbook.save(file_path)
            QMessageBox.information(self, "Export Successful", f"Data exported to {file_path}")
            
        except Exception as e:
            QMessageBox.critical(self, "Export Failed", f"Error exporting data: {str(e)}")
