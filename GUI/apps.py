import sys
from PyQt5 import QtWidgets, uic
from PyQt5.QtCore import Qt
from PyQt5.QtWidgets import QFileDialog, QMessageBox, QTableWidgetItem
from PyQt5.QtGui import QBrush, QColor
import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill
import csv
from lib.EMARK import EMARKPrinter
from lib.IND231 import WeightReader
from lib.PLC import PLCReader
import os
from PyQt5.QtWidgets import QMessageBox
import json
import glob
import time
from collections import deque

CONFIG_FILE = "lib/config.json"

class PrintingSystem(QtWidgets.QMainWindow):
    def __init__(self):
        super(PrintingSystem, self).__init__()

        # Load the UI file
        uic.loadUi('lib/home.ui', self)
        
        # Initialize variables
        self.weight_unit = "lbs"
        self.length_unit = "ft"
        self.weight = 0
        self.length = 0
        self.counter = 1
        self.state_counter = False
        self.counter_text = 1
        self.printed_counter = 0
        self.reject_counter = 0
        self.running = False

        self.pipe_queue = deque()

        self.EMARK = EMARKPrinter()
        self.WEIGHT = WeightReader()
        self.PLC = PLCReader()
        
        # Setup UI elements
        self.setup_ui()

        self.setup_connections()

        self.config = self.load_config()
        self.auto_connect()

        # Get tab widget
        self.tab5_widget = self.findChild(QtWidgets.QWidget, "tab_5")
        self.tab5_index = self.tabWidget.indexOf(self.tab5_widget)
        self.tabWidget.removeTab(self.tab5_index)

        # Create logs directory if it doesn't exist
        if not os.path.exists("logs"):
            os.makedirs("logs")
        
        # Connect signals
        self.connect_signals()
        
    def setup_ui(self):
        # Set initial values
        self.lineEdit_weight.setText("0")
        self.lineEdit_length.setText("0")
        
        # Set default weight and length units
        self.comboBox_weight.setCurrentText("pound (lbs)")
        self.comboBox_length.setCurrentText("foot (ft)")

        # Setup table columns
        self.setup_table()
        self.load_last_csv()

    def setup_table(self):
        self.tableWidget.setColumnCount(11)
        self.tableWidget.setHorizontalHeaderLabels([
            "Date", "Time", "Weight", "Min", "Max", 
            "Length", "Min", "Max", "Counter", 
            "Printed Text", "Status"
        ])
        
        # Set different resize modes for different columns
        header = self.tableWidget.horizontalHeader()
        header.setSectionResizeMode(0, QtWidgets.QHeaderView.ResizeToContents)  # Date
        header.setSectionResizeMode(1, QtWidgets.QHeaderView.ResizeToContents)  # Time
        header.setSectionResizeMode(2, QtWidgets.QHeaderView.ResizeToContents)  # Weight
        header.setSectionResizeMode(3, QtWidgets.QHeaderView.ResizeToContents)  # Min (Weight)
        header.setSectionResizeMode(4, QtWidgets.QHeaderView.ResizeToContents)  # Max (Weight)
        header.setSectionResizeMode(5, QtWidgets.QHeaderView.ResizeToContents)  # Length
        header.setSectionResizeMode(6, QtWidgets.QHeaderView.ResizeToContents)  # Min (Length)
        header.setSectionResizeMode(7, QtWidgets.QHeaderView.ResizeToContents)  # Max (Length)
        header.setSectionResizeMode(8, QtWidgets.QHeaderView.ResizeToContents)  # Counter
        header.setSectionResizeMode(9, QtWidgets.QHeaderView.Stretch)          # Printed Text
        header.setSectionResizeMode(10, QtWidgets.QHeaderView.ResizeToContents) # Status
       
        # Set minimum sizes if needed
        header.setMinimumSectionSize(120)  # Minimum width for all columns

        header = self.tableWidget_2.horizontalHeader()
        header.setSectionResizeMode(QtWidgets.QHeaderView.Stretch)  # Stretch all columns equally
        self.tableWidget_2.setColumnCount(6)
        self.tableWidget_2.setHorizontalHeaderLabels(["Timestamp", "Length", "Weight", "Status Length", "Status Weight", "Status"])

        header = self.tableWidget_3.horizontalHeader()
        header.setSectionResizeMode(QtWidgets.QHeaderView.Stretch)  # Stretch all columns equally
        self.tableWidget_3.setColumnCount(1)
        self.tableWidget_3.setHorizontalHeaderLabels(["Hit Number"])

    def load_config(self):
        if os.path.exists(CONFIG_FILE):
            try:
                with open(CONFIG_FILE, "r") as f:
                    return json.load(f)
            except:
                return {
                    "printer_port": "",
                    "weight_port": "",
                    "plc_ip": "",
                    "min_weight": 0,
                    "max_weight": 100,
                    "min_length": 0,
                    "max_length": 100
                }
        else:
            return {
                "printer_port": "",
                "weight_port": "",
                "plc_ip": "",
                "min_weight": 0,
                "max_weight": 100,
                "min_length": 0,
                "max_length": 100
            }

    def save_config(self):
        try:
            with open(CONFIG_FILE, "w") as f:
                json.dump(self.config, f)
        except:
            pass

    def auto_connect(self):
        # Printer
        for i in range(self.comboBox_com_1.count()):
            if self.config["printer_port"] in self.comboBox_com_1.itemText(i):
                self.comboBox_com_1.setCurrentIndex(i)
                self.connect_printer()
                break

        # Weight
        for i in range(self.comboBox_com_2.count()):
            if self.config["weight_port"] in self.comboBox_com_2.itemText(i):
                self.comboBox_com_2.setCurrentIndex(i)
                self.connect_weight()
                break

        # PLC
        self.lineEdit_IP.setText(self.config["plc_ip"])
        if self.config["plc_ip"]:
            self.connect_PLC()

        self.lineEdit_downlimit.setText(str(self.config["min_weight"]))
        self.lineEdit_uplimit.setText(str(self.config["max_weight"]))
        self.lineEdit_downlimit_1.setText(str(self.config["min_length"]))
        self.lineEdit_uplimit_1.setText(str(self.config["max_length"]))

        self.lineEdit_input.setText(self.config["input_text"])
        self.update_output()

        self.weight_limits_set = True
        self.length_limits_set = True

    def setup_connections(self):
        from serial.tools import list_ports

        ports = list_ports.comports()
        self.comboBox_com_1.clear()
        self.comboBox_com_2.clear()
        for port in ports:
            self.comboBox_com_1.addItem(str(port))
            self.comboBox_com_2.addItem(str(port)) 

    def poll_sensors(self):
        try:
            now = time.time()

            # LENGTH sensor
            length_on = self.PLC.read_bit(byte=5, bit=4)
            if length_on:
                self.length_status.setText("STATUS : MEASURING")
                self.length_status.setStyleSheet("""
                    background-color: rgb(0, 170, 255);
                    border-radius: 5px;
                    border: none;
                """)
                if not hasattr(self, 'length_timer'):
                    self.length_timer = now
                    self.length_processed = False
                elif now - self.length_timer >= 10 and self.length_processed:
                    self.get_length(append = False)
                    self.length_processed = True
                    self.length_timer = now
                elif now - self.length_timer >= 2 and not self.length_processed:
                    self.get_length(append = True)
                    self.length_processed = True
                    self.length_timer = now
            else:
                if self.PLC.connected:
                    self.length_status.setText("STATUS : ONLINE")
                    self.length_status.setStyleSheet("""
                        background-color: rgb(0, 170, 0);
                        border-radius: 5px;
                        border: none;
                    """)
                else:
                    self.length_status.setText("STATUS : DISCONNECTED")
                    self.length_status.setStyleSheet("""
                        background-color: rgb(255, 170, 0);
                        border-radius: 5px;
                        border: none;
                    """)
                self.length_timer = now
                self.length_processed = False

            # WEIGHT sensor
            weight_on = self.PLC.read_bit(byte=4, bit=2)
            if weight_on:
                self.weight_status.setText("STATUS : MEASURING")
                self.weight_status.setStyleSheet("""
                    background-color: rgb(0, 170, 255);
                    border-radius: 5px;
                    border: none;
                """)
                if not hasattr(self, 'weight_timer'):
                    self.weight_timer = now
                    self.weight_processed = False
                elif now - self.weight_timer >= 10  and self.weight_processed:
                    self.get_weight()
                    self.weight_processed = True
                    self.weight_timer = now
                elif now - self.weight_timer >= 2  and not self.weight_processed:
                    self.get_weight()
                    self.weight_processed = True
                    self.weight_timer = now
            else:
                if self.WEIGHT.connected:
                    self.weight_status.setText("STATUS : ONLINE")
                    self.weight_status.setStyleSheet("""
                        background-color: rgb(0, 170, 0);
                        border-radius: 5px;
                        border: none;
                    """)
                else:
                    self.weight_status.setText("STATUS : DISCONNECTED")
                    self.weight_status.setStyleSheet("""
                        background-color: rgb(255, 170, 0);
                        border-radius: 5px;
                        border: none;
                    """)
                self.weight_timer = now
                self.weight_processed = False
            
            # WEIGHT sensor
            trigger_on = self.PLC.read_bit(byte=6, bit=4)
            if trigger_on:
                if not hasattr(self, 'trigger_timer'):
                    self.trigger_timer = now
                    self.trigger_processed = False
                elif now - self.trigger_timer >= 1 and not self.trigger_processed:
                    self.trigger()
                    self.trigger_processed = True
                    self.trigger_timer = now
            else:
                if self.running:
                    self.status_running.setText("STATUS : RUNNING")
                    self.status_running.setStyleSheet("""
                        background-color: rgb(0, 170, 0);
                        border-radius: 5px;
                        border: none;
                    """)
                else:
                    self.status_running.setText("STATUS : OFFLINE")
                    self.status_running.setStyleSheet("""
                        background-color: rgb(255, 170, 0);
                        border-radius: 5px;
                        border: none;
                    """)
                self.trigger_timer = now
                self.trigger_processed = False

        except Exception as e:
            print(f"Sensor read error: {e}")

    def connect_signals(self):
        # Home tab signals
        self.pushButton_weight.clicked.connect(self.apply_weight_limits)
        self.pushButton_length.clicked.connect(self.apply_length_limits)

        self.pushButton_weight_2.clicked.connect(self.get_weight)
        self.pushButton_length_2.clicked.connect(self.get_length_btn)

        self.pushButton_run.clicked.connect(self.run_auto)
        self.pushButton_run_2.clicked.connect(self.trigger)
        self.pushButton_connect_1.clicked.connect(self.connect_printer)
        self.pushButton_connect_2.clicked.connect(self.connect_weight)
        self.pushButton_connect_3.clicked.connect(self.connect_PLC)
        
        # History tab signals
        self.pushButton_open.clicked.connect(self.open_file)
        self.pushButton_save.clicked.connect(self.save_data)
        self.pushButton_export.clicked.connect(self.export_to_excel)
        
        # Combo box changes
        self.comboBox_weight.currentTextChanged.connect(self.update_weight_unit)
        self.comboBox_length.currentTextChanged.connect(self.update_length_unit)

        self.lineEdit_input.textChanged.connect(self.update_output)

        self.checkBox_counter.stateChanged.connect(self.counter_tab)
        self.spinBox_counter.valueChanged.connect(self.handle_spinbox_change)
        self.pushButton_open_counter.clicked.connect(self.open_file_counter)
        
    def counter_tab(self, state):
        if state == 2:  # Checked
            self.state_counter = True
            if self.tabWidget.indexOf(self.tab5_widget) == -1:
                self.tabWidget.addTab(self.tab5_widget, "Counter")
            self.tabWidget.setCurrentWidget(self.tab5_widget)
        else:  # Unchecked
            self.state_counter = False
            index = self.tabWidget.indexOf(self.tab5_widget)
            if index != -1:
                self.tabWidget.removeTab(index)
    
    def handle_spinbox_change(self, value):
        self.counter = self.spinBox_counter.value()
        
        if self.state_counter:
            try:
                self.counter_text = str(self.counter_data[self.counter-1][0])
                self.lineEdit_counter.setText(self.counter_text)
            except Exception as e:
                print("Handle spinBox Error:", e)
                QMessageBox.warning(self, "Invalid Counter Data", "Please Make Sure Counter/Hit Number Data (CSV) Available")
        else:
            self.counter_text = self.counter
            self.lineEdit_counter.setText(f"{self.counter_text:04}")

    def update_weight_unit(self, text):
        """Update the weight unit based on combo box selection"""
        self.weight_unit = text.split("(")[1].strip(")")
        self.label_weight1.setText(f"{self.weight_unit}   -")
        self.label_weight2.setText(self.weight_unit)
    
    def update_length_unit(self, text):
        """Update the length unit based on combo box selection"""
        self.length_unit = text.split("(")[1].strip(")")
        self.label_length1.setText(f"{self.length_unit}   -")
        self.label_length2.setText(self.length_unit)
            
    def apply_weight_limits(self):
        try:
            lower = float(self.lineEdit_downlimit.text())
            upper = float(self.lineEdit_uplimit.text())
            
            if lower >= upper:
                QMessageBox.warning(self, "Invalid Range", "Lower limit must be less than upper limit")
                return
                
            self.weight_limits_set = True

            self.config["min_weight"] = lower
            self.config["max_weight"] = upper
            self.save_config()
            
            QMessageBox.information(self, "Success", "Weight limits applied successfully")
        except ValueError:
            QMessageBox.warning(self, "Invalid Input", "Please enter valid numbers for weight limits")
            
    def apply_length_limits(self):
        try:
            lower = float(self.lineEdit_downlimit_1.text())
            upper = float(self.lineEdit_uplimit_1.text())
            
            if lower >= upper:
                QMessageBox.warning(self, "Invalid Range", "Lower limit must be less than upper limit")
                return
                
            self.length_limits_set = True

            self.config["min_length"] = lower
            self.config["max_length"] = upper
            self.save_config()

            QMessageBox.information(self, "Success", "Length limits applied successfully")
        except ValueError:
            QMessageBox.warning(self, "Invalid Input", "Please enter valid numbers for length limits")
            
    def connect_printer(self):
        if self.EMARK.connected:
            self.EMARK.close()
            self.printer_connected = False
            self.pushButton_connect_1.setText("Connect")
            self.pushButton_connect_1.setStyleSheet("""
                QPushButton{
                    background-color: rgb(255, 170, 0);
                }
                QPushButton:hover{
                    background-color: rgb(255, 190, 0);
                }
                QPushButton:pressed{
                    background-color: rgb(255, 210, 0);
                }
            """)
        else:
            port = self.comboBox_com_1.currentText().split()[0]
            msg = self.EMARK.connect(port= port)
            if self.EMARK.connected:
                self.pushButton_connect_1.setText("Connected")
                self.pushButton_connect_1.setStyleSheet("""
                    QPushButton{
                        background-color: rgb(0, 170, 0);
                    }
                    QPushButton:hover{
                        background-color: rgb(0, 190, 0);
                    }
                    QPushButton:pressed{
                        background-color: rgb(0, 210, 0);
                    }
                """)

                self.config["printer_port"] = port
                self.save_config()
                self.EMARK.clear_text()
            else:
                QMessageBox.critical(self, "Connection Failed", msg)
    
    def connect_weight(self):
        if self.WEIGHT.connected:
            self.WEIGHT.close()
            self.pushButton_connect_2.setText("Connect")
            self.pushButton_connect_2.setStyleSheet("""
                QPushButton{
                    background-color: rgb(255, 170, 0);
                }
                QPushButton:hover{
                    background-color: rgb(255, 190, 0);
                }
                QPushButton:pressed{
                    background-color: rgb(255, 210, 0);
                }
            """)
        else:
            port = self.comboBox_com_2.currentText().split()[0]
            msg = self.WEIGHT.connect(port= port)
            if self.WEIGHT.connected:
                self.pushButton_connect_2.setText("Connected")
                self.pushButton_connect_2.setStyleSheet("""
                    QPushButton{
                        background-color: rgb(0, 170, 0);
                    }
                    QPushButton:hover{
                        background-color: rgb(0, 190, 0);
                    }
                    QPushButton:pressed{
                        background-color: rgb(0, 210, 0);
                    }
                """)
                self.weight_status.setText("STATUS : ONLINE")
                self.weight_status.setStyleSheet("""
                    background-color: rgb(0, 170, 0);
                    border-radius: 5px;
                    border: none;
                """)
                self.config["weight_port"] = port
                self.save_config()
            else:
                QMessageBox.critical(self, "Connection Failed", msg)

    def connect_PLC(self):
        if self.PLC.connected:
            self.PLC.close()
            self.pushButton_connect_3.setText("Connect")
            self.pushButton_connect_3.setStyleSheet("""
                QPushButton{
                    background-color: rgb(255, 170, 0);
                }
                QPushButton:hover{
                    background-color: rgb(255, 190, 0);
                }
                QPushButton:pressed{
                    background-color: rgb(255, 210, 0);
                }
            """)
        else:
            ip_address = self.lineEdit_IP.text()
            msg = self.PLC.connect(ip= ip_address)
            if self.PLC.connected:
                self.pushButton_connect_3.setText("Connected")
                self.pushButton_connect_3.setStyleSheet("""
                    QPushButton{
                        background-color: rgb(0, 170, 0);
                    }
                    QPushButton:hover{
                        background-color: rgb(0, 190, 0);
                    }
                    QPushButton:pressed{
                        background-color: rgb(0, 210, 0);
                    }
                """)
                self.length_status.setText("STATUS : ONLINE")
                self.length_status.setStyleSheet("""
                    background-color: rgb(0, 170, 0);
                    border-radius: 5px;
                    border: none;
                """)

                self.config["plc_ip"] = ip_address
                self.save_config()
            else:
                QMessageBox.critical(self, "Connection Failed", msg)

    def update_output(self):
        input_text = self.lineEdit_input.text()
        output_text = input_text.replace("{WEIGHT}", f"{self.weight} {self.weight_unit.upper()}")
        output_text = output_text.replace("{LENGTH}", f"{self.length} {self.length_unit.upper()}")
        output_text = output_text.replace("{COUNTER}", f"{self.counter}")
        
        self.lineEdit_output.setText(output_text)

        self.config["input_text"] = input_text
        self.save_config()

    def check_setup(self):
        if not self.EMARK.connected:
            QMessageBox.warning(self, "Printer Not Connected", "Please connect to the printer first")
            return
        if not self.WEIGHT.connected:
            QMessageBox.warning(self, "Weight Indicator Not Connected", "Please connect to Weight Indicator first")
            return
        if not self.PLC.connected:
            QMessageBox.warning(self, "PLC Not Connected", "Please connect to the PLC first")
            return

    def run_auto(self):
        if not self.running:
            # Start polling
            self.sensor_timer = QTimer()
            self.sensor_timer.timeout.connect(self.poll_sensors)
            self.sensor_timer.start(500)

            self.running = True
            self.pushButton_run.setText("STOP")
            self.pushButton_run.setStyleSheet("""
                QPushButton {
                    background-color: rgb(255, 170, 0);
                    border-radius: 10px;
                }
                QPushButton:hover {
                    background-color: rgb(255, 190, 0);
                }
                QPushButton:pressed {
                    background-color: rgb(255, 210, 0);
                }
            """)

            self.status_running.setText("STATUS : RUNNING")
            self.status_running.setStyleSheet("""
                background-color: rgb(0, 170, 0);
                border-radius: 5px;
                border: none;
            """)
        else:
            # Stop polling
            if hasattr(self, 'sensor_timer'):
                self.sensor_timer.stop()
                self.sensor_timer.deleteLater()
            
            self.running = False
            self.pushButton_run.setText("RUN")
            self.pushButton_run.setStyleSheet("""
                QPushButton{
                    background-color: rgb(0, 170, 0);
                    border-radius: 10px;
                }
                QPushButton:hover{
                    background-color: rgb(0, 190, 0);
                }
                QPushButton:pressed{
                    background-color: rgb(0, 210, 0);
                }
            """)

            self.status_running.setText("STATUS : OFFLINE")
            self.status_running.setStyleSheet("""
                background-color: rgb(255, 170, 0);
                border-radius: 5px;
                border: none;
            """)
    
    def get_length_btn(self):
        self.get_length(append=True)
                
    def get_length(self, append = True):
        try:
            length = round(self.PLC.read_real(db_number=2, start_byte=0), 2)

            if self.length_unit == "ft":
                length /= 304.8

            self.lineEdit_length.setText(f"{length:.2f}")

            # Determine status
            status_length = "OK"
            status = "WAITING FOR PRINTING"
            if length < self.config["min_length"]:
                status_length = "UNDERLENGTH"
                status = "REJECT"
            elif length > self.config["max_length"]:
                status_length = "OVERLENGTH"
                status = "REJECT"

            pipe_data = {
                'timestamp': datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                'length': f"{length:.2f}",
                'weight': None,
                'status_length': status_length,
                'status_weight': None,
                'status': status,
            }
            if append:
                if len(self.pipe_queue)>0:
                    if self.pipe_queue[-1]['weight'] == None:
                        self.pipe_queue[-1]['timestamp'] = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                        self.pipe_queue[-1]['length'] = f"{length:.2f}"
                        self.pipe_queue[-1]['status_length'] = status_length
                        self.pipe_queue[-1]['status'] = status
                    else: 
                        self.pipe_queue.append(pipe_data)
                else: 
                    self.pipe_queue.append(pipe_data)
            else:
                if len(self.pipe_queue)>0:
                    self.pipe_queue[-1]['timestamp'] = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    self.pipe_queue[-1]['length'] = f"{length:.2f}"
                    self.pipe_queue[-1]['status_length'] = status_length
                    self.pipe_queue[-1]['status'] = status

            self.update_queue_table()
            
            if status_length != "OK":
                QMessageBox.warning(self, "Quality Alert", f"Pipe rejected: {status_length}")

        except Exception as e:
            print(e)
            QMessageBox.warning(self, "Sensor Alert", f"PLC Not Connected or Error!")
    
    def get_weight(self):
        try:
            raw_weight = self.WEIGHT.read_weight()

            if raw_weight is not None:
                weight = round(raw_weight, 2)

                if self.weight_unit == "lbs":
                    weight /= 2.205
                    
                self.lineEdit_weight.setText(f"{weight:.2f}")

                # Determine status
                status_weight = "OK"
                status = "WAITING FOR PRINTING"
                if weight < self.config["min_weight"]:
                    status_weight = "UNDERWEIGHT"
                    status = "REJECT"
                elif weight > self.config["max_weight"]:
                    status_weight = "OVERWEIGHT"
                    status = "REJECT"

                if len(self.pipe_queue)>0:
                    self.pipe_queue[-1]["weight"] = f"{weight:.2f}"
                    self.pipe_queue[-1]["status_weight"] = status_weight
                    self.pipe_queue[-1]["status"] = status
                    self.update_queue_table()

                if status_weight != "OK":
                    QMessageBox.warning(self, "Quality Alert", f"Pipe rejected: {status_weight}")

            else:
                QMessageBox.warning(self, "Sensor Alert", f"Weight Indicator Not Connected or Error!")
        except Exception as e:
            print(e)
            QMessageBox.warning(self, "Sensor Alert", f"Weight Indicator Not Connected or Error!")

    def trigger(self):
        try:
            self.status_running.setText("STATUS : SENDING DATA TO PRINTER")
            self.status_running.setStyleSheet("""
                background-color: rgb(0, 170, 255);
                border-radius: 5px;
                border: none;
            """)

            oldest_pipe = self.pipe_queue.popleft()

            length = oldest_pipe['length']
            weight = oldest_pipe['weight']
            status = oldest_pipe['status']

            # Process input text
            input_text = self.lineEdit_input.text()
            output_text = input_text.replace("{WEIGHT}", f"{weight} {self.weight_unit.upper()}")
            output_text = output_text.replace("{LENGTH}", f"{length} {self.length_unit.upper()}")
            output_text = output_text.replace("{COUNTER}", f"{self.counter_text}")
            self.lineEdit_output.setText(output_text)
            print(output_text)

            if status == "WAITING FOR PRINTING":
                status = "OK"
                response = self.EMARK.send_text(output_text, 
                                        template_num=1, 
                                        font_height=0x0C,  # 16 dot matrix
                                        x_pos=0, 
                                        y_pos=0)
                if response:
                    print(f"Response: {response}")
                else:
                    QMessageBox.warning(self, "Hardware Error", f"EMark Printer Not Connected or Error!")
                
                self.printed_counter += 1
                self.lcdNumber_printed.display(self.printed_counter)
            elif status == "REJECT":
                self.reject_counter+=1
                self.lcdNumber_reject.display(self.reject_counter)
                self.EMARK.clear_text()

            # Add to history (whether printed or rejected)
            self.add_to_history(weight, length, self.counter_text, output_text, status)
            self.update_queue_table()

            self.counter+=1
            self.spinBox_counter.setValue(self.counter)
            
        except Exception as e:
            print(e)
            QMessageBox.critical(self, "Error", f"An error occurred: {str(e)}")

    def update_queue_table(self):
        self.lcdNumber_queue.display(len(self.pipe_queue))
        self.tableWidget_2.setRowCount(len(self.pipe_queue))

        for i, item in enumerate(reversed(self.pipe_queue)):
            items = [
                QTableWidgetItem(item["timestamp"]),
                QTableWidgetItem(item["length"]),
                QTableWidgetItem(item["weight"]),
                QTableWidgetItem(item["status_length"]),
                QTableWidgetItem(item["status_weight"]),
                QTableWidgetItem(item["status"])
            ]

            for j, cell in enumerate(items):
                cell.setTextAlignment(Qt.AlignCenter)
                self.tableWidget_2.setItem(i, j, cell)

    def add_to_history(self, weight, length, counter, output_text, status):
        current_time = datetime.datetime.now()
        date_str = current_time.strftime("%Y-%m-%d")
        time_str = current_time.strftime("%H:%M:%S")
        
        row_position = self.tableWidget.rowCount()
        self.tableWidget.insertRow(row_position)
        
        weight_lower = self.lineEdit_downlimit.text()
        weight_upper = self.lineEdit_uplimit.text()
        length_lower = self.lineEdit_downlimit_1.text()
        length_upper = self.lineEdit_uplimit_1.text()
        
        for col, value in enumerate([
            date_str,
            time_str,
            f"{weight} {self.weight_unit}",
            f"{weight_lower}",
            f"{weight_upper}",
            f"{length} {self.length_unit}",
            f"{length_lower}",
            f"{length_upper}",
            str(counter),
            output_text,
            status
        ]):
            item = QTableWidgetItem(value)
            item.setTextAlignment(Qt.AlignCenter)
            item.setFlags(item.flags() ^ Qt.ItemIsEditable)
            
            # Color code based on status
            if col == 10 and status != "OK":  # Status column is now index 10
                if "WEIGHT" in status:
                    item.setBackground(QBrush(QColor("yellow")))
                elif "LENGTH" in status:
                    item.setBackground(QBrush(QColor("orange")))
            
            self.tableWidget.setItem(row_position, col, item)

        # Save to CSV
        self.save_to_csv(date_str, time_str, weight, length, counter, output_text, status)

    def save_to_csv(self, date, time, weight, length, counter, output_text, status):
        current_time = datetime.datetime.now()
        date_str = current_time.strftime("%Y-%m-%d")
        csv_file = f"logs/{date_str}.csv"
        
        file_exists = os.path.isfile(csv_file)
        
        weight_min = self.lineEdit_downlimit.text()
        weight_max = self.lineEdit_uplimit.text()
        length_min = self.lineEdit_downlimit_1.text()
        length_max = self.lineEdit_uplimit_1.text()
        
        try:
            with open(csv_file, 'a', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                
                if not file_exists:
                    writer.writerow([
                        "Date", "Time", "Weight", "Min", "Max", 
                        "Length", "Min", "Max", "Counter", 
                        "Printed Text", "Status"
                    ])
                
                writer.writerow([
                    date,
                    time,
                    f"{weight} {self.weight_unit}",
                    weight_min,
                    weight_max,
                    f"{length} {self.length_unit}",
                    length_min,
                    length_max,
                    str(counter),
                    output_text,
                    status
                ])
        except Exception as e:
            print(f"Error saving to CSV: {e}")

    def load_last_csv(self):
        try:
            csv_files = glob.glob("logs/*.csv")
            if not csv_files:
                return

            latest_file = max(csv_files, key=os.path.getmtime)
            self.lineEdit_path.setText(latest_file)

            with open(latest_file, 'r', newline='', encoding='utf-8') as f:
                reader = csv.reader(f)
                headers = next(reader)  # Skip header row

                self.tableWidget.setRowCount(0)

                for row_data in reader:
                    row_position = self.tableWidget.rowCount()
                    self.tableWidget.insertRow(row_position)
                    for col, value in enumerate(row_data):
                        item = QTableWidgetItem(value)
                        item.setTextAlignment(Qt.AlignCenter)
                        item.setFlags(item.flags() ^ Qt.ItemIsEditable)
                        
                        # Apply color coding for rejected items when loading
                        if col == 10 and value != "OK":  # Status column is now index 10
                            if "WEIGHT" in value:
                                item.setBackground(QBrush(QColor("yellow")))
                            elif "LENGTH" in value:
                                item.setBackground(QBrush(QColor("orange")))
                                
                        self.tableWidget.setItem(row_position, col, item)
        except Exception as e:
            print(f"Failed to load last CSV: {e}")

    def open_file(self):
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "Open CSV File",
            "./logs",  # or use absolute path like "/home/user/logs"
            "CSV Files (*.csv)"
        )
        if file_path:
            self.lineEdit_path.setText(file_path)
            try:
                with open(file_path, 'r', newline='', encoding='utf-8') as f:
                    reader = csv.reader(f)
                    headers = next(reader)

                    self.tableWidget.setRowCount(0)

                    for row_data in reader:
                        row_position = self.tableWidget.rowCount()
                        self.tableWidget.insertRow(row_position)
                        for col, value in enumerate(row_data):
                            item = QTableWidgetItem(value)
                            item.setTextAlignment(Qt.AlignCenter)
                            item.setFlags(item.flags() ^ Qt.ItemIsEditable)
                            
                            # Apply color coding for rejected items
                            if col == 10 and value != "OK":  # Status column
                                if "WEIGHT" in value:
                                    item.setBackground(QBrush(QColor("yellow")))
                                elif "LENGTH" in value:
                                    item.setBackground(QBrush(QColor("orange")))
                                    
                            self.tableWidget.setItem(row_position, col, item)
            except Exception as e:
                QMessageBox.critical(self, "Error", f"Failed to open file: {str(e)}")
    
    def save_data(self):
        QMessageBox.information(self, "Data Saved", "Data has been saved successfully")

    def open_file_counter(self):
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "Open CSV File",
            "./counter",
            "CSV Files (*.csv)"
        )
        if file_path:
            self.lineEdit_path_2.setText(file_path)
            try:
                with open(file_path, 'r', newline='', encoding='utf-8') as f:
                    reader = csv.reader(f)
                    self.counter_data = list(reader)
                    print("Loaded Data:", self.counter_data)
                    self.counter = 1
                    self.spinBox_counter.setValue(self.counter)
                    self.counter_text = str(self.counter_data[self.counter-1][0])
                    self.lineEdit_counter.setText(self.counter_text)

                    self.tableWidget_3.setRowCount(0)

                    for row_data in self.counter_data:
                        row_position = self.tableWidget_3.rowCount()
                        self.tableWidget_3.insertRow(row_position)
                        for col, value in enumerate(row_data):
                            item = QTableWidgetItem(value)
                            item.setTextAlignment(Qt.AlignCenter)
                            item.setFlags(item.flags() ^ Qt.ItemIsEditable)
                            self.tableWidget_3.setItem(row_position, col, item)

            except Exception as e:
                QMessageBox.critical(self, "Error", f"Failed to open file: {str(e)}")
  
    def export_to_excel(self):
        if self.tableWidget.rowCount() == 0:
            QMessageBox.warning(self, "No Data", "There is no data to export")
            return
            
        file_path, _ = QFileDialog.getSaveFileName(self, "Export to Excel", "", "Excel Files (*.xlsx);;All Files (*)")
        if file_path:
            if not file_path.endswith('.xlsx'):
                file_path += '.xlsx'
                
            try:
                workbook = openpyxl.Workbook()
                sheet = workbook.active
                sheet.title = "Printing History"
                
                # Write headers
                headers = []
                for col in range(self.tableWidget.columnCount()):
                    headers.append(self.tableWidget.horizontalHeaderItem(col).text())
                sheet.append(headers)
                
                # Make headers bold
                for cell in sheet[1]:
                    cell.font = Font(bold=True)
                
                # Write data with color coding
                for row in range(self.tableWidget.rowCount()):
                    row_data = []
                    for col in range(self.tableWidget.columnCount()):
                        item = self.tableWidget.item(row, col)
                        row_data.append(item.text() if item else "")
                    sheet.append(row_data)
                    
                    # Apply color to entire row based on status
                    status_item = self.tableWidget.item(row, 10)  # Status column is now index 10
                    if status_item and status_item.text() != "OK":
                        fill_color = "FFFF00" if "WEIGHT" in status_item.text() else "FFA500"
                        for cell in sheet[row+2]:  # +2 because headers are row 1 and Excel is 1-based
                            cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                
                workbook.save(file_path)
                QMessageBox.information(self, "Export Successful", f"Data exported to {file_path}")
                
            except Exception as e:
                QMessageBox.critical(self, "Export Failed", f"Error exporting data: {str(e)}")

if __name__ == "__main__":
    from PyQt5.QtGui import QPixmap
    from PyQt5.QtWidgets import QSplashScreen
    from PyQt5.QtCore import QTimer

    app = QtWidgets.QApplication(sys.argv)

    # Create splash screen
    splash_pix = QPixmap("lib/logo.jpg")  # <-- you can use any image here
    splash = QSplashScreen(splash_pix)
    splash.show()
    app.processEvents()  # Show the splash screen immediately

    window = PrintingSystem()
    window.show()
    splash.finish(window)

    sys.exit(app.exec_())
